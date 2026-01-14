from datetime import datetime
import io
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlmodel import Session, select
from fastapi.responses import Response
from sqlalchemy import func, or_
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from urllib.parse import quote
from app.db import get_session
from app.schemas import ToolCreate, ToolRead, ToolListResponse,MovementAction
from app.schemas import ToolQuantityUpdate
from app.schemas import ToolListItem
from app.deps import require_user
from app.models import Tool, User, ToolMovement
from app.services.ledger import calc_signed_delta_and_new_qty, build_note, abort

router = APIRouter(prefix="/tools", tags=["tools"])

@router.post("", response_model=ToolRead)
def create_tool(
        data: ToolCreate,
        session: Session = Depends(get_session),
        _user: User = Depends(require_user),
):
    tool = Tool(
        name=data.name,
        location=data.location,
        quantity=data.quantity,
    )
    session.add(tool)
    session.flush()  # 生成 tool.id

    if tool.quantity > 0:
        mv = ToolMovement(
            tool_id=tool.id,
            action=MovementAction.IN.value,  # ✅ 统一口径
            delta=tool.quantity,
            note="新建入库",
            operator=_user.username,
        )
        tool.updated_at = datetime.utcnow()
        session.add(mv)

    session.commit()
    session.refresh(tool)
    return tool


@router.get("", response_model=ToolListResponse)
def list_tools(
        q: str | None = None,
        limit: int = Query(50, ge=1, le=200),
        offset: int = Query(0, ge=0),
        sort: str = Query(
            "id_desc",
            description="排序：id_desc/id_asc/name_asc/name_desc/qty_asc/qty_desc",
        ),
        session: Session = Depends(get_session),
        _user: User = Depends(require_user),
):
    conds = []
    if q:
        conds.append(or_(Tool.name.contains(q), Tool.location.contains(q)))

    # total
    count_stmt = select(func.count()).select_from(Tool)
    if conds:
        count_stmt = count_stmt.where(*conds)
    total = session.exec(count_stmt).one()

    # order by
    order_map = {
        "id_desc": Tool.id.desc(),
        "id_asc": Tool.id.asc(),
        "name_asc": Tool.name.asc(),
        "name_desc": Tool.name.desc(),
        "qty_asc": Tool.quantity.asc(),
        "qty_desc": Tool.quantity.desc(),
    }
    if sort not in order_map:
        abort(400, "BAD_REQUEST", f"sort 不支持：{sort}")
    order_by = order_map[sort]

    # items
    items_stmt = select(Tool)
    if conds:
        items_stmt = items_stmt.where(*conds)

    items = session.exec(items_stmt.order_by(order_by).offset(offset).limit(limit)).all()

    return {
        "items": items,
        "total": total,
        "limit": limit,
        "offset": offset,
        "q": q,
    }


@router.get("/lite", response_model=list[ToolListItem])
def list_tools_lite(
        session: Session = Depends(get_session),
        _user: User = Depends(require_user),
):
    stmt = select(Tool).order_by(Tool.id.desc())
    return session.exec(stmt).all()


@router.get("/export.xlsx")
def export_tools_xlsx(
    q: str | None = None,
    session: Session = Depends(get_session),
    _user: User = Depends(require_user),
):
    stmt = select(Tool).order_by(Tool.id.asc())
    if q:
        stmt = stmt.where(
            or_(
                Tool.name.contains(q),
                Tool.location.contains(q),
            )
        )
    tools = session.exec(stmt).all()

    header_cn = ["编号", "名称", "库位", "数量", "品牌", "型号", "备注", "更新时间"]

    def norm_str(v, default: str) -> str:
        if v is None:
            return default
        s = str(v).strip()
        return s if s else default

    def norm_int(v, default: int = 0) -> int:
        if v is None:
            return default
        try:
            return int(v)
        except Exception:
            return default

    def norm_dt_obj(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.replace(tzinfo=None) if v.tzinfo else v
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "刀具台账"

    # 表头样式（依旧保留，Table 也会有样式，但这让首行更“台账”）
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.append(header_cn)
    ws.row_dimensions[1].height = 26  # 表头行高度（可改成 24/26 更“厚”）
    for col in range(1, len(header_cn) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 数据行
    for t in tools:
        d = t.model_dump()
        ws.append([
            norm_int(d.get("id"), 0),
            norm_str(d.get("name"), "未命名"),
            norm_str(d.get("location"), "未知"),
            norm_int(d.get("quantity"), 0),
            norm_str(d.get("vendor"), ""),
            norm_str(d.get("model"), ""),
            norm_str(d.get("remark"), ""),
            norm_dt_obj(d.get("updated_at")),
        ])

    data_end_row = 1 + len(tools)  # 表头+数据

    # ✅ 冻结首行
    ws.freeze_panes = "A2"

    # ✅ 列格式：数量/更新时间
    for r in range(2, data_end_row + 1):
        ws.cell(row=r, column=4).number_format = "0"  # 数量
        ws.cell(row=r, column=8).number_format = "yyyy-mm-dd hh:mm:ss"  # 更新时间

    # ✅ 列宽（稳定台账风格）
    col_widths = {
        "A": 8,   # 编号
        "B": 22,  # 名称
        "C": 12,  # 库位
        "D": 8,   # 数量
        "E": 12,  # 品牌
        "F": 16,  # 型号
        "G": 28,  # 备注
        "H": 20,  # 更新时间
    }
    for k, w in col_widths.items():
        ws.column_dimensions[k].width = w

    # ✅ 加 Table 样式（只覆盖表头+数据）
    # 如果没有数据，也至少给到表头行，避免范围非法
    last_row = max(1, data_end_row)
    table_ref = f"A1:H{last_row}"

    table = Table(displayName=f"ToolsLedger_{datetime.now().strftime('%H%M%S')}", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",   # 你也可以换成 Medium2/Medium10 等
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,        # ✅ 条纹行
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # 末尾：导出时间（不在 Table 范围里）
    ws.append([])
    ws.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    cn_filename = "刀具台账.xlsx"
    quoted = quote(cn_filename)
    headers = {
        "Content-Disposition": f"attachment; filename=\"tools.xlsx\"; filename*=UTF-8''{quoted}"
    }

    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.patch("/{tool_id}/quantity", response_model=ToolRead)
def update_tool_quantity(
    tool_id: int,
    body: ToolQuantityUpdate,
    session: Session = Depends(get_session),
    user: User = Depends(require_user),
):
    tool = session.get(Tool, tool_id)
    if not tool:
        abort(404, "NOT_FOUND", "Tool not found")

    old_qty = tool.quantity
    signed_delta, new_qty = calc_signed_delta_and_new_qty(body.action, body.delta, old_qty)

    tool.quantity = new_qty
    tool.updated_at = datetime.utcnow()

    note = build_note(body.action, body.delta, old_qty, new_qty, body.note)

    mv = ToolMovement(
        tool_id=tool.id,
        action=body.action.value,
        delta=signed_delta,
        note=note,
        operator=user.username,
    )

    session.add(tool)
    session.add(mv)
    session.commit()
    session.refresh(tool)
    return tool


@router.delete("/{tool_id}")
def delete_tool(
        tool_id: int,
        session: Session = Depends(get_session),
        _user: User = Depends(require_user),
):
    tool = session.get(Tool, tool_id)
    # 查询资源：session.get(Tool, tool_id)
    # 通过 tool_id 从数据库中查询 Tool 模型对应的记录（get 方法按主键查询，效率高于 filter）
    if not tool:
        abort(404, "NOT_FOUND", "Tool not found")
    session.delete(tool)
    session.commit()
    return {"ok": True}


@router.get("/{tool_id}", response_model=ToolRead)
def get_tool(
        tool_id: int,
        session: Session = Depends(get_session),
        _user: User = Depends(require_user),
):
    tool = session.get(Tool, tool_id)
    if not tool:
        abort(404, "NOT_FOUND", "Tool not found")
    return tool